"""
Parse DataRealis Excel files for La Flèche (72154)
- 72154_resultats_elections.xlsx: 9 elections, 17 bureaux each
- 72154_liste_electorale_statistiques.xlsx: demographics per bureau

Output: data/historique_bv_enriched.json
"""
import openpyxl
import json
import math
import os

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
RAW_DIR = os.path.join(BASE_DIR, "raw")
OUT_FILE = os.path.join(BASE_DIR, "historique_bv.json")

# ── Election config: sheet name → normalized key + bloc grouping ──
ELECTIONS = [
    {
        "sheet": "Européennes 2019",
        "key": "europeennes2019",
        "date": "2019-05-26",
        "type": "europeennes",
        "blocs": {
            "RN": ["RN"],
            "Gauche": ["EELV", "PS", "LFI"],
            "Macron": ["LREM"],
            "LR": ["LR"],
            "Reconquete": ["DLF"],  # closest equivalent in 2019
            "Autres": []  # catch-all
        }
    },
    {
        "sheet": "Municipales 2020 T1",
        "key": "municipales2020",
        "date": "2020-03-15",
        "type": "municipales",
        "blocs": None  # raw candidate names
    },
    {
        "sheet": "Présidentielle 2022 T1",
        "key": "presidentielle2022t1",
        "date": "2022-04-10",
        "type": "presidentielle",
        "blocs": {
            "RN": ["RN"],
            "Gauche": ["LFI", "EELV"],
            "Macron": ["LREM"],
            "LR": ["LR", "LASSALLE"],
            "Reconquete": ["REC"],
            "Autres": []
        }
    },
    {
        "sheet": "Présidentielle 2022 T2",
        "key": "presidentielle2022t2",
        "date": "2022-04-24",
        "type": "presidentielle",
        "blocs": None  # only 2 candidates
    },
    {
        "sheet": "Législatives 2022 T1",
        "key": "legislatives2022t1",
        "date": "2022-06-12",
        "type": "legislatives",
        "blocs": {
            "RN": ["RN"],
            "Gauche": ["NUP", "DVG"],
            "Macron": ["ENS"],
            "LR": ["LR", "DSV"],
            "Reconquete": ["REC"],
            "Autres": []
        }
    },
    {
        "sheet": "Législatives 2022 T2",
        "key": "legislatives2022t2",
        "date": "2022-06-19",
        "type": "legislatives",
        "blocs": None  # only 2 candidates
    },
    {
        "sheet": "Européennes 2024",
        "key": "europeennes2024",
        "date": "2024-06-09",
        "type": "europeennes",
        "blocs": {
            "RN": ["RN"],
            "Gauche": ["PS", "LFI", "EELV", "COM"],
            "Macron": ["ENS"],
            "LR": ["LR"],
            "Reconquete": ["REC"],
            "Autres": []
        }
    },
    {
        "sheet": "Législatives 2024 T1",
        "key": "legislatives2024t1",
        "date": "2024-07-07",
        "type": "legislatives",
        "blocs": {
            "RN": ["RN"],
            "Gauche": ["UG"],
            "Macron": ["ENS"],
            "LR": ["LR"],
            "Reconquete": ["REC"],
            "Autres": []
        }
    },
    {
        "sheet": "Législatives 2024 T2",
        "key": "legislatives2024t2",
        "date": "2024-07-07",
        "type": "legislatives",
        "blocs": None  # only 2 candidates
    },
]


def parse_election_sheet(ws, election_cfg):
    """Parse one sheet into structured data."""
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]
    data_rows = rows[1:]

    # Identify vote columns (those without "en %" suffix and not metadata)
    meta_cols = {"Bureau", "Inscrits", "Votants", "Abstentions", "Exprimés"}
    candidate_cols = []
    for i, h in enumerate(headers):
        if h and h not in meta_cols and "(en %)" not in str(h):
            candidate_cols.append((i, h))

    result = {
        "date": election_cfg["date"],
        "type": election_cfg["type"],
        "candidats": [name for _, name in candidate_cols],
        "bureaux": {},
        "commune": {}
    }

    # If bloc grouping is defined, compute blocs
    bloc_cfg = election_cfg.get("blocs")

    for row in data_rows:
        bureau_name = row[0]
        if not bureau_name:
            continue

        inscrits = row[1] or 0
        votants = row[2] or 0
        exprimes = row[4] or 0

        # Raw votes per candidate
        voix = {}
        for col_idx, cand_name in candidate_cols:
            voix[cand_name] = row[col_idx] or 0

        entry = {
            "inscrits": inscrits,
            "votants": votants,
            "exprimes": exprimes,
            "voix": voix
        }

        # Compute blocs if defined
        if bloc_cfg:
            blocs_voix = {}
            assigned = set()
            for bloc_name, parties in bloc_cfg.items():
                if bloc_name == "Autres":
                    continue
                total = sum(voix.get(p, 0) for p in parties)
                blocs_voix[bloc_name] = total
                assigned.update(parties)

            # Autres = everything not assigned
            autres = sum(v for k, v in voix.items() if k not in assigned)
            blocs_voix["Autres"] = autres
            entry["blocs"] = blocs_voix

        if bureau_name == "Commune":
            result["commune"] = entry
        else:
            # Extract bureau number from "La FlècheXX"
            bv_num = "".join(c for c in bureau_name if c.isdigit())
            if bv_num.startswith("0"):
                bv_num = bv_num.lstrip("0") or "0"
            result["bureaux"][bv_num] = entry

    return result


def parse_demographics(filepath):
    """Parse liste électorale statistics."""
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb["Statistiques"]
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]

    demographics = {}
    for row in rows[1:]:
        bureau_name = row[0]
        if not bureau_name or bureau_name == "Commune":
            continue

        bv_num = "".join(c for c in str(bureau_name) if c.isdigit())
        if bv_num.startswith("0"):
            bv_num = bv_num.lstrip("0") or "0"

        demographics[bv_num] = {
            "inscrits": row[1] or 0,
            "nationalite": {
                "francaise": row[2] or 0,
                "portugaise": row[3] or 0,
                "belge": row[4] or 0,
                "espagnole": row[5] or 0,
                "italienne": row[6] or 0,
                "irlandaise": row[7] or 0,
            },
            "age": {
                "18_35": row[8] or 0,
                "36_65": row[9] or 0,
                "66_80": row[10] or 0,
                "80_plus": row[11] or 0,
            }
        }

    wb.close()
    return demographics


def compute_volatility(elections_data):
    """
    Compute per-bureau volatility from multi-election data.
    For each bureau, calculate the std dev of each bloc's % across elections.
    This gives data-driven sigmas for Monte Carlo.
    """
    # Only use elections with bloc groupings (T1 elections)
    bloc_elections = {k: v for k, v in elections_data.items()
                      if "blocs" in (v.get("commune") or {})}

    if not bloc_elections:
        return {}

    # Get list of all bureau IDs
    all_bv = set()
    for el in bloc_elections.values():
        all_bv.update(el["bureaux"].keys())

    volatility = {}
    for bv in sorted(all_bv, key=lambda x: int(x)):
        bloc_pcts = {}  # bloc -> list of percentages across elections
        for el_key, el_data in bloc_elections.items():
            if bv not in el_data["bureaux"]:
                continue
            bureau = el_data["bureaux"][bv]
            if "blocs" not in bureau or bureau["exprimes"] == 0:
                continue
            for bloc, voix in bureau["blocs"].items():
                pct = (voix / bureau["exprimes"]) * 100
                bloc_pcts.setdefault(bloc, []).append(pct)

        # Compute std dev for each bloc
        bv_vol = {}
        for bloc, pcts in bloc_pcts.items():
            if len(pcts) >= 2:
                mean = sum(pcts) / len(pcts)
                variance = sum((p - mean) ** 2 for p in pcts) / (len(pcts) - 1)
                bv_vol[bloc] = round(math.sqrt(variance), 2)
            else:
                bv_vol[bloc] = 5.0  # default if only 1 election
        volatility[bv] = bv_vol

    return volatility


def compute_commune_volatility(elections_data):
    """Compute commune-level volatility for each bloc."""
    bloc_elections = {k: v for k, v in elections_data.items()
                      if "blocs" in (v.get("commune") or {})}

    bloc_pcts = {}
    for el_key, el_data in bloc_elections.items():
        commune = el_data["commune"]
        if commune["exprimes"] == 0:
            continue
        for bloc, voix in commune["blocs"].items():
            pct = (voix / commune["exprimes"]) * 100
            bloc_pcts.setdefault(bloc, []).append(pct)

    volatility = {}
    for bloc, pcts in bloc_pcts.items():
        if len(pcts) >= 2:
            mean = sum(pcts) / len(pcts)
            variance = sum((p - mean) ** 2 for p in pcts) / (len(pcts) - 1)
            volatility[bloc] = round(math.sqrt(variance), 2)
        else:
            volatility[bloc] = 5.0
    return volatility


def main():
    elections_file = os.path.join(RAW_DIR, "72154_resultats_elections.xlsx")
    demographics_file = os.path.join(RAW_DIR, "72154_liste_electorale_statistiques.xlsx")

    # Parse elections
    wb = openpyxl.load_workbook(elections_file, read_only=True)
    elections_data = {}
    for cfg in ELECTIONS:
        ws = wb[cfg["sheet"]]
        elections_data[cfg["key"]] = parse_election_sheet(ws, cfg)
        nb = len(elections_data[cfg["key"]]["bureaux"])
        print(f"  ✓ {cfg['key']}: {nb} bureaux")
    wb.close()

    # Parse demographics
    demographics = parse_demographics(demographics_file)
    print(f"  ✓ Demographics: {len(demographics)} bureaux")

    # Compute volatility
    volatility = compute_volatility(elections_data)
    commune_vol = compute_commune_volatility(elections_data)
    print(f"  ✓ Volatility computed for {len(volatility)} bureaux")
    print(f"    Commune volatility: {commune_vol}")

    # Build output
    output = {
        "commune": "La Flèche",
        "codeInsee": "72154",
        "codePostal": "72200",
        "nbBureaux": 17,
        "source": "DataRealis + data.gouv.fr",
        "elections": elections_data,
        "demographics": demographics,
        "volatility": {
            "bureaux": volatility,
            "commune": commune_vol,
            "description": "Ecart-type (%) de chaque bloc sur les elections T1 avec blocs (EUR2019, PRES2022T1, LEG2022T1, EUR2024, LEG2024T1)"
        }
    }

    with open(OUT_FILE, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    size_kb = os.path.getsize(OUT_FILE) / 1024
    print(f"\n✅ Output: {OUT_FILE} ({size_kb:.1f} KB)")


if __name__ == "__main__":
    main()
